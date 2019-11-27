# -*- coding: utf-8 -*-

"""
.. module:: generate_cim.py
   :license: GPL/CeCIL
   :platform: Unix, Windows
   :synopsis: Generates CMIP6 JSON documents from XLS files.

.. moduleauthor:: Mark Conway-Greenslade <momipsl@ipsl.jussieu.fr>

"""
import argparse
import collections
import json
import os

import openpyxl
import pyessv

import _utils as utils
from cmip6.utils import logger
from cmip6.utils import vocabs



# Define command line argument parser.
_ARGS = argparse.ArgumentParser("Generates CMIP6 model JSON files.")
_ARGS.add_argument(
    "--institution-id",
    help="An institution identifier",
    dest="institution_id",
    type=str
    )

# MIP era.
_MIP_ERA = "cmip6"

# Name of file controlling publication.
_MODEL_PUBLICATION_FNAME = "model_publication.json"


def _main(args):
    """Main entry point.

    """
    institutes = vocabs.get_institutes(args.institution_id)
    for i in institutes:
        for s in vocabs.get_institute_sources(i):
            for t in pyessv.ESDOC.cmip6.get_model_topics(s):
                try:
                    wb = _get_spreadsheet(i, s, t)
                except IOError:
                    warning = '{} :: {} :: {} spreadsheet not found'
                    warning = warning.format(i.canonical_name, s.canonical_name, t.canonical_name)
                    logger.log_warning(warning)
                    continue

                _write_to_fs(i, s, t, _get_content(i, s, t, wb))


def _get_spreadsheet(i, s, t):
    """Returns a model topic spreadsheet for processing.

    """
    fname = utils.get_file_of_cmip6(i, s, t, 'xlsx')
    path = utils.get_folder_of_cmip6_source(i, s)
    path = os.path.join(path, fname)
    if not os.path.exists:
        raise IOError()

    return openpyxl.load_workbook(path, read_only=True)


def _get_content(i, s, t, wb):
    """Returns content to be written to file system.

    """
    # Initialise output.
    obj = collections.OrderedDict()
    obj['mipEra'] = _MIP_ERA
    obj['institute'] = i.canonical_name
    obj['seedingSource'] = 'Spreadsheet'
    obj['sourceID'] = s.canonical_name
    obj['topic'] = t.canonical_name
    obj['content'] = collections.OrderedDict()

    # Process spreadsheet.
    for idx, ws in enumerate(wb):
        # Process citations/responsible parties.
        if idx == 1:
            # TODO
            continue

        # Extract specialization entries.
        elif idx > 1:
            _set_xls_content(obj, ws)

    return obj


def _write_to_fs(i, s, t, obj):
    """Writes json content to file system.

    """
    fname = utils.get_file_of_cmip6(i, s, t, 'json')
    path = utils.get_folder_of_cmip6_source(i, s, 'json')
    path = os.path.join(path, fname)

    # Write only when there is content.
    if len(obj['content']) > 0:
        with open(path, 'w') as fstream:
            fstream.write(json.dumps(obj, indent=4))


def _set_xls_content(obj, ws):
    """Sets content for a particular specialization.

    """
    content = None
    content_is_enum = None
    for row in ws.iter_rows(min_row=1, max_col=4, max_row=ws.max_row):
        # Separation row - begin new specialization block.
        if row[1].value is None:
            if content is not None:
                if content[1]:
                    obj['content'][content[0]] = {'values': content[1]}
                content = None

        # Specialization begin row - spec. id is hidden in 3rd column.
        elif row[2].value is not None:
            content = (row[2].value, [])
            content_is_enum = row[0].value == 'ENUM'

        # Specialization value row.
        elif content is not None and not _is_note(row[1]):
            value = row[1].value

            # Open enums are treated differently.
            if content_is_enum:
                if row[3].value:
                    value = 'Other: ' + row[3].value
                elif value == 'Other: document in cell to the right':
                    break

            # This can occur when spreadsheet is deformatted.
            if value == '=TRUE()':
                value = True
            elif value == '=FALSE()':
                value = False

            # If value is not in blocklist then emit.
            if value not in ('-', 'Other: -'):
                content[1].append(value)


def _is_note(cell):
    """Returns flag indicating whether a cell represents a note tot he user or not.

    """
    try:
        return cell.value.startswith('NOTE: ')
    except AttributeError:
        return False


# Main entry point.
if __name__ == '__main__':
    _main(_ARGS.parse_args())
