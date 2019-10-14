# -*- coding: utf-8 -*-

"""
.. module:: generate_cim.py
   :license: GPL/CeCIL
   :platform: Unix, Windows
   :synopsis: Generates CMIP6 JSON documents from XLS files.

.. moduleauthor:: Mark Conway-Greenslade <momipsl@ipsl.jussieu.fr>

"""
import argparse
import collections
import json
import os

import openpyxl
import pyessv

import _utils as utils



# Define command line argument parser.
_ARGS = argparse.ArgumentParser("Generates CMIP6 responsible parties JSON files.")
_ARGS.add_argument(
    "--institution-id",
    help="An institution identifier",
    dest="institution_id",
    type=str
    )

# MIP era.
_MIP_ERA = "cmip6"


def _main(args):
    """Main entry point.

    """
    # Set institutes to be processed.
    institutes = pyessv.WCRP.cmip6.institution_id if args.institution_id == 'all' else \
                 [pyessv.WCRP.cmip6.institution_id[args.institution_id]]

    # Write a JSON file per CMIP6 institute | source | topic combination.
    # i = institute | s = source | t = topic
    for i in institutes:
        try:
            wb = _get_spreadsheet(i)
        except IOError:
            warning = '{} responsible parties spreadsheet not found'
            warning = warning.format(i.canonical_name)
            pyessv.log_warning(warning)
            continue

        _write_to_fs(i, _get_content(i, wb))


def _get_spreadsheet(i):
    """Returns a spreadsheet for processing.

    """
    fname = 'cmip6_{}_responsible_parties.xlsx'.format(i.canonical_name)
    path = utils.get_folder((i, 'cmip6', 'responsible_parties'))
    path = os.path.join(path, fname)
    if not os.path.exists:
        raise IOError()

    return openpyxl.load_workbook(path, read_only=True)


def _get_content(i, wb):
    """Returns content to be written to file system.

    """
    # Initialise output.
    obj = collections.OrderedDict()
    obj['mipEra'] = _MIP_ERA
    obj['institute'] = i.canonical_name
    obj['seedingSource'] = 'Spreadsheet'
    obj['content'] = []

    # Process spreadsheet.
    for idx, ws in enumerate(wb):
        if idx > 1:
            _set_xls_content(obj, ws)

    return obj


def _write_to_fs(i, obj):
    """Writes json content to file system.

    """
    fname = 'cmip6_{}_responsible_parties.json'.format(i.canonical_name)
    path = utils.get_folder((i, 'cmip6', 'responsible_parties', 'json'))
    if not os.path.exists(path):
        os.makedirs(path)
    path = os.path.join(path, fname)

    # # Write only when there is content.
    if len(obj['content']) > 0:
        with open(path, 'w') as fstream:
            fstream.write(json.dumps(obj, indent=4))


def _set_xls_content(obj, ws):
    """Sets content for a particular specialization.

    """
    content = None
    for row in ws.iter_rows(min_row=3, max_col=7, max_row=ws.max_row):
        if row[0].value is None:
            continue

        party = collections.OrderedDict()
        party['mnemonic'], \
        party['name'], \
        party['is_organisation'], \
        party['address_postal'], \
        party['address_email'], \
        party['url'], \
        party['orcid'] = (i.value for i in row)
        party['is_organisation'] = True if party['is_organisation'] in ['yes', 'y'] else False
        obj['content'].append(party)


# Main entry point.
if __name__ == '__main__':
    _main(_ARGS.parse_args())
