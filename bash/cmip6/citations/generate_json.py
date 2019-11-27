# -*- coding: utf-8 -*-

"""
.. module:: generate_json.py
   :license: GPL/CeCIL
   :platform: Unix, Windows
   :synopsis: Generates CMIP6 JSON documents from XLS files.

.. moduleauthor:: Mark Conway-Greenslade <momipsl@ipsl.jussieu.fr>

"""
import argparse
import collections
import json
import os

import openpyxl

from cmip6.utils import io_mgr
from cmip6.utils import logger
from cmip6.utils import vocabs


# Define command line argument parser.
_ARGS = argparse.ArgumentParser("Generates CMIP6 citation JSON files.")
_ARGS.add_argument(
    "--institution-id",
    help="An institution identifier",
    dest="institution_id",
    type=str
    )


def _main(args):
    """Main entry point.

    """
    for i in vocabs.get_institutes(args.institution_id):
        _write(i)


def _write(i):
    """Writes citation JSON file for a particular institute.

    """
    try:
        spreadsheet = _get_spreadsheet(i)
    except IOError:
        msg = '{} citations spreadsheet not found'.format(i.canonical_name)
        logger.log_warning(msg)
    else:
        content = _get_content(i, spreadsheet)
        if content:
            _write_content(i, content)


def _get_spreadsheet(i):
    """Returns a spreadsheet for processing.

    """
    path = io_mgr.get_citations_spreadsheet(i)
    if not os.path.exists(path):
        raise IOError()

    return openpyxl.load_workbook(path, read_only=True)


def _get_content(i, spreadsheet):
    """Returns content to be written to file system.

    """
    # Initialise output.
    obj = collections.OrderedDict()
    obj['mipEra'] = "cmip6"
    obj['institute'] = i.canonical_name
    obj['seedingSource'] = 'Spreadsheet'
    obj['content'] = []

    # Process spreadsheet.
    for _, worksheet in [(i, j) for i, j in enumerate(spreadsheet) if i > 1]:
        _set_xls_content(obj, worksheet)

    return obj


def _write_content(i, content):
    """Writes json content to file system.

    """
    fpath = io_mgr.get_citations_json(i)
    with open(fpath, 'w') as fstream:
        fstream.write(json.dumps(content, indent=4))


def _set_xls_content(obj, worksheet):
    """Sets content for a particular specialization.

    """
    for row in worksheet.iter_rows(min_row=3, max_col=5, max_row=worksheet.max_row):
        mnemonic, doi, bibtex, url, long_name = [i.value for i in row]
        if mnemonic is None:
            continue
        if doi is None and bibtex is None and url is None and long_name is None:
            continue

        citation = collections.OrderedDict()
        citation['mnemonic'] = mnemonic
        citation['doi'] = doi
        citation['bibtex'] = bibtex
        citation['url'] = url
        citation['long_name'] = long_name

        obj['content'].append(citation)


# Main entry point.
if __name__ == '__main__':
    _main(_ARGS.parse_args())
